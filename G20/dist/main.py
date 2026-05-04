import os
import sys
import time
import json
import requests
import serial
import os
import re
import time
import random
import requests
from urllib.parse import urljoin
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.tianqihoubao.com"
START_URL = "https://www.tianqihoubao.com/lishi/yunnan.htm"
OUTPUT_DIR = "太仓历史天气"


cookies = {
    "Hm_lvt_d26809c29069303b20c4de5135f3ff06": "1777297322",
    "HMACCOUNT": "CD9B77F951A7ACAC",
    "ASP.NET_SessionId": "xmooqcdiqhtp2fgfr5wzfzix",
    "Hm_lpvt_d26809c29069303b20c4de5135f3ff06": "1777301695",
}

headers = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-language": "zh-CN,zh;q=0.9",
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "priority": "u=0, i",
    "referer": "https://www.tianqihoubao.com/lishi/kunming/month/202601.html",
    "sec-ch-ua": '"Not:A-Brand";v="99", "Google Chrome";v="145", "Chromium";v="145"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "same-origin",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
}


session = requests.Session()
session.headers.update(headers)
session.cookies.update(cookies)


def safe_filename(name: str) -> str:
    """
    清理 Windows 文件名非法字符
    """
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    name = name.strip()
    return name or "未命名城市"


def text_clean(value: str) -> str:
    """
    清理多余空白
    """
    if value is None:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def sleep_random():
    """
    随机延迟，避免请求过快
    """
    time.sleep(random.uniform(0.5, 1.2))


def fetch_html(url: str, retry: int = 3):
    """
    请求页面并返回 etree.HTML 对象
    """
    last_error = None

    for i in range(retry):
        try:
            response = session.get(url, timeout=15)
            response.raise_for_status()

            # 自动识别中文编码
            if not response.encoding or response.encoding.lower() == "iso-8859-1":
                response.encoding = response.apparent_encoding

            return etree.HTML(response.text)

        except Exception as e:
            last_error = e
            print(f"[请求失败] 第 {i + 1}/{retry} 次: {url} -> {e}")
            time.sleep(1 + i)

    raise RuntimeError(f"请求失败: {url}, error={last_error}")


def get_city_links():
    """
    从云南城市列表页提取城市链接

    XPath:
    /html/body/div/div[2]/div[1]/div[2]

    目标：
    每个 dl 下的 dt 下的 a 标签 href 和 title
    """
    tree = fetch_html(START_URL)

    city_nodes = tree.xpath("/html/body/div/div[2]/div[1]/div[2]//dl/dt/a")

    cities = []
    seen = set()

    for a in city_nodes:
        href = text_clean(a.get("href"))
        title = text_clean(a.get("title") or "".join(a.xpath(".//text()")))

        if not href or not title:
            continue

        city_url = urljoin(BASE_URL, href)

        key = (title, city_url)
        if key in seen:
            continue

        seen.add(key)

        cities.append({
            "title": title,
            "href": href,
            "url": city_url,
        })

    return cities


def get_month_links(city_url: str):
    """
    访问城市历史天气页，提取第一个和第二个 class 包含 card-body 的 div 下所有 a 标签 href

    例如：
    https://www.tianqihoubao.com/lishi/kunming.html
    """
    tree = fetch_html(city_url)

    # 取第一个和第二个 class 包含 card-body 的 div
    card_bodies = tree.xpath(
        '(//div[contains(concat(" ", normalize-space(@class), " "), " card-body ")])[position() <= 2]'
    )

    months = []
    seen = set()

    for div in card_bodies:
        a_nodes = div.xpath(".//a[@href]")

        for a in a_nodes:
            href = text_clean(a.get("href"))
            title = text_clean(a.get("title") or "".join(a.xpath(".//text()")))

            if not href:
                continue

            month_url = urljoin(BASE_URL, href)

            if month_url in seen:
                continue

            seen.add(month_url)

            months.append({
                "title": title,
                "href": href,
                "url": month_url,
            })

    return months


def parse_month_weather(month_url: str):
    """
    解析某个月份页面的 tbody 数据

    页面结构类似：

    <tbody>
        <tr>
            <td><a>2026年01月01日</a></td>
            <td>晴 / 多云</td>
            <td>16℃ / 4℃</td>
            <td>北风 1-3级 / 北风 1-3级</td>
        </tr>
    </tbody>

    只取四列：
    日期
    天气状况(白天/夜间)
    最高/最低气温
    风力风向(白天/夜间)
    """
    tree = fetch_html(month_url)

    trs = tree.xpath("//tbody/tr")

    rows = []

    for tr in trs:
        tds = tr.xpath("./td")

        # 跳过分隔线、空行、广告行等
        if len(tds) < 4:
            continue

        # 第一列必须是日期链接
        date_a = tds[0].xpath(".//a")
        if not date_a:
            continue

        date_text = text_clean("".join(tds[0].xpath(".//text()")))
        weather = text_clean("".join(tds[1].xpath(".//text()")))
        temperature = text_clean("".join(tds[2].xpath(".//text()")))
        wind = text_clean("".join(tds[3].xpath(".//text()")))

        if not date_text:
            continue

        rows.append([
            date_text,
            weather,
            temperature,
            wind,
        ])

    return rows


def create_city_xlsx(city_title: str, rows: list):
    """
    每个城市创建一个 xlsx 文件
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    filename = safe_filename(city_title) + ".xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "历史天气"

    header = [
        "日期",
        "天气状况(白天/夜间)",
        "最高/最低气温",
        "风力风向(白天/夜间)",
    ]

    ws.append(header)

    for row in rows:
        ws.append(row)

    # 表头样式
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 内容样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    # 设置列宽
    widths = {
        1: 18,
        2: 28,
        3: 18,
        4: 34,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)

    print(f"[保存完成] {filepath}，共 {len(rows)} 条数据")


def main():
    cities = get_city_links()

    print(f"[城市数量] {len(cities)}")

    for city_index, city in enumerate(cities, start=1):
        city_title = city["title"]
        city_url = city["url"]

        print(f"\n[{city_index}/{len(cities)}] 开始处理城市：{city_title}")
        print(f"城市链接：{city_url}")

        all_rows = []

        try:
            sleep_random()

            month_links = get_month_links(city_url)

            print(f"[{city_title}] 找到月份链接数量：{len(month_links)}")

            for month_index, month in enumerate(month_links, start=1):
                month_title = month["title"]
                month_url = month["url"]

                print(f"  [{month_index}/{len(month_links)}] 抓取月份：{month_title}")
                print(f"  月份链接：{month_url}")

                try:
                    sleep_random()

                    rows = parse_month_weather(month_url)

                    print(f"    获取天气数据：{len(rows)} 条")

                    all_rows.extend(rows)

                except Exception as e:
                    print(f"    [月份失败] {city_title} - {month_title} -> {e}")

            create_city_xlsx(city_title, all_rows)

        except Exception as e:
            print(f"[城市失败] {city_title} -> {e}")

def get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = get_base_dir()

CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
WEATHER_FILE = os.path.join(BASE_DIR, "weather.txt")


def load_config():
    """
    读取配置文件 config.json
    """
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config = json.load(f)

        amap_key = config.get("amap_key")
        city_code = config.get("city_code")
        port = config.get("port")
        baudrate = int(config.get("baudrate", 115200))

        if not amap_key or not city_code or not port:
            raise ValueError("config.json 配置不完整")

        return {
            "amap_key": amap_key,
            "city_code": city_code,
            "port": port,
            "baudrate": baudrate
        }

    except Exception as e:
        print("读取 config.json 失败：", e)
        return None


def fetch_weather(amap_key, city_code):
    url = "https://restapi.amap.com/v3/weather/weatherInfo"

    params = {
        "key": amap_key,
        "city": city_code
    }

    try:
        res = requests.get(url, params=params, timeout=5)
        data = res.json()

        live = data["lives"][0]

        temperature = str(live["temperature"])
        humidity = str(live["humidity"])

        return temperature, humidity

    except Exception as e:
        print("天气抓取失败：", e)
        return None


def save_weather(temp, hum):
    with open(WEATHER_FILE, "w", encoding="utf-8") as f:
        f.write(str(temp).strip() + "\n")
        f.write(str(hum).strip() + "\n")


def read_weather():
    try:
        with open(WEATHER_FILE, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()

        if len(lines) < 2:
            return None

        temp = lines[0].strip()
        hum = lines[1].strip()

        if not temp or not hum:
            return None

        return temp, hum

    except Exception as e:
        print("读取 weather.txt 失败：", e)
        return None


def send_to_m5(port, baudrate, temp, hum):
    text = f"temp={temp},hum={hum}\n"

    try:
        with serial.Serial(port, baudrate, timeout=2) as ser:
            time.sleep(2)
            ser.write(text.encode("utf-8"))
            ser.flush()

        print("已发送：", text.strip())
        print("串口：", port)
        print("波特率：", baudrate)

        return True

    except Exception as e:
        print("串口发送失败：", e)
        return False


def main():
    config = load_config()

    if not config:
        print("程序退出")
        return

    amap_key = config["amap_key"]
    city_code = config["city_code"]
    port = config["port"]
    baudrate = config["baudrate"]

    weather = fetch_weather(amap_key, city_code)

    if weather:
        temp, hum = weather
        save_weather(temp, hum)
        print("天气抓取成功，已保存 weather.txt")
    else:
        weather = read_weather()

        if not weather:
            print("没有可用的历史天气数据，程序退出")
            return

        temp, hum = weather
        print("使用上一次 weather.txt 数据")

    send_to_m5(port, baudrate, temp, hum)

if __name__ == "__main__":
    main()